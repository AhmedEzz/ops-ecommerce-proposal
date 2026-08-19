#!/usr/bin/env python3
"""
Build the Phase 1 budget model for the OPS Management System.

Everything downstream of the Assumptions sheet is a formula, so changing a rate,
an FTE split or the project window recalculates the whole workbook. Contingency
is deliberately the balancing figure: it is computed as
budget - labour - infrastructure, which guarantees the total lands on exactly the
client's 300,000 EGP whatever the inputs are changed to.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "deliverables" / "02-Budget-Plan.xlsx"

FONT = "Arial"

# Financial-model colour convention: blue = hardcoded input, black = formula,
# green = link to another sheet, yellow fill = assumption the client must confirm.
INPUT = Font(name=FONT, size=10, color="0000FF")
FORMULA = Font(name=FONT, size=10, color="000000")
LINK = Font(name=FONT, size=10, color="008000")
BODY = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
TITLE = Font(name=FONT, size=14, bold=True, color="1F3864")
H2 = Font(name=FONT, size=11, bold=True, color="FFFFFF")
NOTE = Font(name=FONT, size=9, italic=True, color="595959")

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
BAND_FILL = PatternFill("solid", fgColor="D9E2F3")
FLAG_FILL = PatternFill("solid", fgColor="FFFF00")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
GOOD_FILL = PatternFill("solid", fgColor="E2EFDA")

EGP = '#,##0;(#,##0);-'
EGP2 = '#,##0.00;(#,##0.00);-'
PCT = '0.0%'
NUM = '0.00'
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def title(ws, text, subtitle=None, width=8):
    ws["A1"] = text
    ws["A1"].font = TITLE
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = NOTE
    ws.freeze_panes = "A4"


def header(ws, row, labels, start=1):
    for i, label in enumerate(labels):
        cell = ws.cell(row=row, column=start + i, value=label)
        cell.font = H2
        cell.fill = HEAD_FILL
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def put(ws, ref, value, font=BODY, fmt=None, fill=None, border=True, align=None):
    cell = ws[ref]
    cell.value = value
    cell.font = font
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    if border:
        cell.border = BOX
    if align:
        cell.alignment = Alignment(horizontal=align)
    return cell


wb = Workbook()

# ---------------------------------------------------------------- README ----
ws = wb.active
ws.title = "README"
widths(ws, {"A": 30, "B": 92})
title(ws, "OPS Management System — Phase 1 Budget Model",
      "Client budget: 300,000 EGP · Window: 10 Aug – 20 Nov 2026 · Team: BE + FE + DevOps")

rows = [
    ("Purpose", "Models the Phase 1 build against a fixed 300,000 EGP client budget, and states "
                "explicitly what that budget does and does not buy."),
    ("How to use it", "Edit only the blue cells on the Assumptions sheet. Every other figure in "
                      "the workbook is a formula and will recalculate."),
    ("Colour legend", "BLUE text = input you can change · BLACK = calculated · GREEN = pulled "
                      "from another sheet · YELLOW fill = assumption the client must confirm."),
    ("Balancing figure", "Contingency is calculated as budget − labour − infrastructure. The "
                         "grand total therefore always equals the budget exactly."),
    ("", ""),
    ("Headline", "The budget funds ~8.3 person-months. That covers the Phase 1 scope in the "
                 "technical proposal and nothing beyond it."),
    ("Key constraint", "Bringing returns in-house costs ~3 person-weeks more than the E-stebdal "
                       "integration it replaces. This was absorbed by moving the customer "
                       "self-service portal to Phase 1b — which leaves Phase 1 with NO slack."),
    ("Implication", "Any scope added after Sprint 0 must displace something else. See the "
                    "Sensitivity sheet for the three levers available."),
    ("Scope change", "Authentication and the Super Admin console were added after the first "
                     "draft. They cost ~10 days inside Phase 1, paid for by moving three "
                     "read-only surfaces to Phase 1b. See the Scope Changes sheet."),
    ("", ""),
    ("Sheets", "Assumptions · Resource Allocation · Cost Breakdown · Infra and Tooling · "
               "Cash Flow · Scope Changes · Phase 1b and 2 · Sensitivity · ROI"),
    ("Source", "Figures derive from POC - Youssey Excelsheet.xlsx via scripts/extract_data.py "
               "and from published provider pricing (Aug 2026). Salary bands: Egyptian "
               "mid-level market, 18,000–42,000 EGP/month."),
]
r = 4
for label, text in rows:
    if label:
        put(ws, f"A{r}", label, BOLD, border=False)
        c = put(ws, f"B{r}", text, BODY, border=False)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 30 if len(text) > 95 else 15
    r += 1

# ----------------------------------------------------------- ASSUMPTIONS ----
ws = wb.create_sheet("Assumptions")
widths(ws, {"A": 42, "B": 16, "C": 12, "D": 62})
title(ws, "Assumptions", "Edit the BLUE cells only. Yellow fill = confirm with the client.")

header(ws, 4, ["Assumption", "Value", "Unit", "Basis / source"])

A = [
    ("Client budget (Phase 1)", 300000, "EGP", "Client-stated budget", INPUT, EGP, FLAG_FILL),
    ("Project start", "2026-08-10", "date", "Sprint 0 begins", INPUT, None, None),
    ("Code freeze / go-live", "2026-11-20", "date", "7 days before Black Friday (27 Nov 2026)", INPUT, None, None),
    ("Black Friday", "2026-11-27", "date", "Fixed external date", INPUT, None, None),
    ("Project duration", 15, "weeks", "10 Aug – 20 Nov 2026", INPUT, NUM, None),
    ("Weeks per month", 4.345, "weeks", "52 / 12", INPUT, NUM, None),
    ("", None, None, None, None, None, None),
    ("USD / EGP", 50.25, "EGP", "Market rate, Aug 2026", INPUT, EGP2, FLAG_FILL),
    ("EUR / EGP", 55.00, "EGP", "Derived from USD/EGP at EURUSD ~1.09", INPUT, EGP2, FLAG_FILL),
    ("", None, None, None, None, None, None),
    ("Backend engineer — monthly cost", 34000, "EGP", "Mid-level Egyptian market; the critical role", INPUT, EGP, FLAG_FILL),
    ("Frontend engineer — monthly cost", 27000, "EGP", "Mid-level; Livewire/Blade profile, not React-only", INPUT, EGP, FLAG_FILL),
    ("DevOps engineer — monthly cost", 35000, "EGP", "Mid-level; part-time allocation", INPUT, EGP, FLAG_FILL),
    ("", None, None, None, None, None, None),
    ("Backend allocation", 1.00, "FTE", "Full time across all sprints", INPUT, NUM, None),
    ("Frontend allocation", 1.00, "FTE", "Full time across all sprints", INPUT, NUM, None),
    ("DevOps allocation", 0.40, "FTE", "Front-loaded Sprint 1, then part time", INPUT, NUM, FLAG_FILL),
    ("", None, None, None, None, None, None),
    ("Infrastructure months funded", 5, "months", "Aug–Dec 2026, covering hypercare", INPUT, NUM, None),
]
r = 5
for label, value, unit, basis, font, fmt, fill in A:
    if label:
        put(ws, f"A{r}", label, BOLD)
        put(ws, f"B{r}", value, font or INPUT, fmt, fill)
        put(ws, f"C{r}", unit, BODY, align="center")
        c = put(ws, f"D{r}", basis, NOTE)
        c.alignment = Alignment(wrap_text=True, vertical="center")
    r += 1

put(ws, "A25", "Duration in months", BOLD, fill=BAND_FILL)
put(ws, "B25", "=B9/B10", FORMULA, NUM, BAND_FILL)
put(ws, "C25", "months", BODY, fill=BAND_FILL, align="center")
put(ws, "D25", "Project duration ÷ weeks per month", NOTE, fill=BAND_FILL)

put(ws, "A26", "Total person-months", BOLD, fill=BAND_FILL)
put(ws, "B26", "=(B19+B20+B21)*B25", FORMULA, NUM, BAND_FILL)
put(ws, "C26", "person-mo", BODY, fill=BAND_FILL, align="center")
put(ws, "D26", "Sum of FTE allocations × duration in months", NOTE, fill=BAND_FILL)

put(ws, "A28", "NOTE", BOLD, border=False)
c = put(ws, "B28",
        "Team model is IN-HOUSE SALARIED. Contractor rates typically run 1.3–1.8× and agency "
        "rates 2–3× these figures — at either, 300,000 EGP would not fund the Phase 1 scope. "
        "See the Sensitivity sheet.", NOTE, border=False)
c.alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("B28:D29")

# ---------------------------------------------------- RESOURCE ALLOCATION ----
ws = wb.create_sheet("Resource Allocation")
widths(ws, {"A": 10, "B": 20, "C": 34, "D": 10, "E": 10, "F": 10, "G": 14})
title(ws, "Resource Allocation by Sprint",
      "FTE per role per sprint. Sprint effort is indicative; cost is driven by the Assumptions sheet.")

header(ws, 4, ["Sprint", "Dates", "Primary focus", "BE FTE", "FE FTE", "Ops FTE", "Person-weeks"])

SPRINTS = [
    ("0", "10–14 Aug", "Discovery, credentials, schema, CS return-policy workshop", 1.0, 0.5, 1.0, 1),
    ("1", "17–28 Aug", "Schema, RBAC, audit log, webhook inbox, Filament shell", 1.0, 1.0, 0.8, 2),
    ("2", "31 Aug–11 Sep", "Shopify adapter, order ingestion, item-level splitting", 1.0, 1.0, 0.3, 2),
    ("3", "14–25 Sep", "Bosta adapter, per-item AWB, status webhooks, flex-ship", 1.0, 1.0, 0.3, 2),
    ("4", "28 Sep–9 Oct", "Financial engine: snapshots, ledger, 3 business models", 1.0, 1.0, 0.3, 2),
    ("5", "12–23 Oct", "Returns & exchanges in-house, eligibility, reverse pickup", 1.0, 1.0, 0.4, 2),
    ("6", "26 Oct–6 Nov", "Settlement, reconciliation, vendor portal, migration", 1.0, 1.0, 0.3, 2),
    ("UAT", "9–20 Nov", "UAT, parallel run, load test, cutover, hypercare prep", 1.0, 1.0, 0.6, 2),
]
r = 5
for sprint, dates, focus, be, fe, ops, weeks in SPRINTS:
    put(ws, f"A{r}", sprint, BOLD, align="center")
    put(ws, f"B{r}", dates, BODY)
    c = put(ws, f"C{r}", focus, BODY)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    put(ws, f"D{r}", be, INPUT, NUM, align="center")
    put(ws, f"E{r}", fe, INPUT, NUM, align="center")
    put(ws, f"F{r}", ops, INPUT, NUM, align="center")
    put(ws, f"G{r}", f"=(D{r}+E{r}+F{r})*{weeks}", FORMULA, NUM, align="center")
    r += 1

put(ws, "C13", "Total person-weeks", BOLD, fill=BAND_FILL, align="right")
put(ws, "D13", "", BODY, fill=BAND_FILL)
put(ws, "E13", "", BODY, fill=BAND_FILL)
put(ws, "F13", "", BODY, fill=BAND_FILL)
put(ws, "G13", "=SUM(G5:G12)", BOLD, NUM, BAND_FILL, align="center")

put(ws, "A15", "Why DevOps is 0.4 FTE", BOLD, border=False)
c = put(ws, "B15",
        "Effort is front-loaded into Sprint 1 (environments, CI/CD, secrets, backups) and Sprint 5 "
        "(load test), with light steady-state work between. On a two-node deployment there is no "
        "Kubernetes or managed-broker work to justify a full-time role, and the hours are better "
        "spent on integrations — which is where the delivery risk actually sits.", NOTE, border=False)
c.alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("B15:G17")

# ---------------------------------------------------------- COST BREAKDOWN ----
ws = wb.create_sheet("Cost Breakdown")
widths(ws, {"A": 34, "B": 14, "C": 12, "D": 14, "E": 16, "F": 12, "G": 46})
title(ws, "Cost Breakdown — Phase 1",
      "Contingency is the balancing figure, so the grand total always equals the client budget exactly.")

header(ws, 4, ["Line item", "Monthly (EGP)", "FTE", "Months", "Total (EGP)", "% of budget", "Notes"])

put(ws, "A5", "LABOUR", BOLD, fill=BAND_FILL)
for col in "BCDEFG":
    put(ws, f"{col}5", "", BODY, fill=BAND_FILL)

LAB = [
    ("Backend engineer", "Assumptions!$B$15", "Assumptions!$B$19",
     "Owns integrations, item splitting, financial engine, returns module"),
    ("Frontend engineer", "Assumptions!$B$16", "Assumptions!$B$20",
     "Filament theming, dashboards, vendor portal, print layouts, Phase 1b portal"),
    ("DevOps engineer", "Assumptions!$B$17", "Assumptions!$B$21",
     "Environments, CI/CD, monitoring, backups, load test, cutover"),
]
r = 6
for label, rate_ref, fte_ref, note in LAB:
    put(ws, f"A{r}", label, BODY)
    put(ws, f"B{r}", f"={rate_ref}", LINK, EGP)
    put(ws, f"C{r}", f"={fte_ref}", LINK, NUM, align="center")
    put(ws, f"D{r}", "=Assumptions!$B$25", LINK, NUM, align="center")
    put(ws, f"E{r}", f"=B{r}*C{r}*D{r}", FORMULA, EGP)
    put(ws, f"F{r}", f"=E{r}/Assumptions!$B$5", FORMULA, PCT, align="center")
    c = put(ws, f"G{r}", note, NOTE)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    r += 1

put(ws, "A9", "Labour subtotal", BOLD, fill=GOOD_FILL)
for col in "BCD":
    put(ws, f"{col}9", "", BODY, fill=GOOD_FILL)
put(ws, "E9", "=SUM(E6:E8)", BOLD, EGP, GOOD_FILL)
put(ws, "F9", "=E9/Assumptions!$B$5", BOLD, PCT, GOOD_FILL, align="center")
put(ws, "G9", "", BODY, fill=GOOD_FILL)

put(ws, "A11", "NON-LABOUR", BOLD, fill=BAND_FILL)
for col in "BCDEFG":
    put(ws, f"{col}11", "", BODY, fill=BAND_FILL)

put(ws, "A12", "Infrastructure and tooling", BODY)
put(ws, "B12", "='Infra and Tooling'!$E$18", LINK, EGP)
put(ws, "C12", "", BODY)
put(ws, "D12", "=Assumptions!$B$23", LINK, NUM, align="center")
put(ws, "E12", "='Infra and Tooling'!$F$18", LINK, EGP)
put(ws, "F12", "=E12/Assumptions!$B$5", FORMULA, PCT, align="center")
c = put(ws, "G12", "Hosting, storage, CI, email, domain — itemised on the Infra sheet", NOTE)
c.alignment = Alignment(wrap_text=True, vertical="center")

put(ws, "A13", "Contingency", BODY, fill=WARN_FILL)
for col in "BCD":
    put(ws, f"{col}13", "", BODY, fill=WARN_FILL)
put(ws, "E13", "=Assumptions!$B$5-E9-E12", FORMULA, EGP, WARN_FILL)
put(ws, "F13", "=E13/Assumptions!$B$5", FORMULA, PCT, WARN_FILL, align="center")
c = put(ws, "G13", "BALANCING FIGURE = budget − labour − infrastructure. Covers integration "
                   "surprises and the PDPL hosting fallback (see Sensitivity).", NOTE, fill=WARN_FILL)
c.alignment = Alignment(wrap_text=True, vertical="center")

put(ws, "A15", "GRAND TOTAL", BOLD, fill=HEAD_FILL)
ws["A15"].font = H2
for col in "BCD":
    put(ws, f"{col}15", "", BODY, fill=HEAD_FILL)
c = put(ws, "E15", "=E9+E12+E13", BOLD, EGP, HEAD_FILL)
c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
c = put(ws, "F15", "=E15/Assumptions!$B$5", BOLD, PCT, HEAD_FILL, align="center")
c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
put(ws, "G15", "", BODY, fill=HEAD_FILL)

put(ws, "A17", "Check — total equals budget", BOLD)
put(ws, "B17", '=IF(ROUND(E15-Assumptions!$B$5,2)=0,"OK — balances exactly","MISMATCH")',
    FORMULA, None, GOOD_FILL)
ws.merge_cells("B17:C17")

put(ws, "A19", "Cost per person-month", BOLD)
put(ws, "B19", "=E9/Assumptions!$B$26", FORMULA, EGP)
c = put(ws, "D19", "Against the Egyptian mid-level band of 18,000–42,000 EGP/month, this sits "
                   "mid-range — achievable, but not a rate that buys senior engineers.", NOTE, border=False)
c.alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("D19:G20")

# ------------------------------------------------------- INFRA AND TOOLING ----
ws = wb.create_sheet("Infra and Tooling")
widths(ws, {"A": 30, "B": 26, "C": 12, "D": 10, "E": 16, "F": 16, "G": 40})
title(ws, "Infrastructure and Tooling",
      "Recommended: Hetzner Cloud CX line (EU) fronted by Cloudflare. See proposal §8.")

header(ws, 4, ["Item", "Spec", "Native cost", "Currency", "EGP / month", "Total (EGP)", "Notes"])

ITEMS = [
    ("Application node", "CX43 — 8 vCPU / 16 GB", 15.99, "EUR", "Ops console, vendor portal, workers, scheduler"),
    ("Data node", "CX33 — 4 vCPU / 8 GB", 8.49, "EUR", "PostgreSQL 16 + Redis 7"),
    ("Staging", "CX23 — 2 vCPU / 4 GB", 5.49, "EUR", "Full stack, sandbox integration credentials"),
    ("Automated backups", "20% of server cost", 5.99, "EUR", "Nightly snapshots, 30-day retention"),
    ("Object storage", "Storage Box 1 TB", 5.00, "EUR", "AWB PDFs, payout proofs, off-site backups"),
    ("Source control / CI", "GitHub Team, 3 seats", 12.00, "USD", "Actions free tier is sufficient"),
    ("Transactional email", "Amazon SES", 5.00, "USD", "AWB labels and payout notifications to vendors"),
    ("DNS / TLS / WAF / CDN", "Cloudflare free tier", 0.00, "USD", "Also provides DDoS protection"),
    ("Error tracking", "Sentry free tier", 0.00, "USD", "5k events/month covers this volume"),
    ("Uptime monitoring", "Uptime Kuma (self-hosted)", 0.00, "USD", "Runs on the staging node"),
]
r = 5
for label, spec, cost, cur, note in ITEMS:
    put(ws, f"A{r}", label, BODY)
    put(ws, f"B{r}", spec, BODY)
    put(ws, f"C{r}", cost, INPUT, EGP2, align="center")
    put(ws, f"D{r}", cur, BODY, align="center")
    fx = "Assumptions!$B$13" if cur == "EUR" else "Assumptions!$B$12"
    put(ws, f"E{r}", f"=C{r}*{fx}", FORMULA, EGP)
    put(ws, f"F{r}", f"=E{r}*Assumptions!$B$23", FORMULA, EGP)
    c = put(ws, f"G{r}", note, NOTE)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    r += 1

put(ws, "A15", "Domain registration", BODY)
put(ws, "B15", "Annual, amortised", BODY)
put(ws, "C15", 600, INPUT, EGP, align="center")
put(ws, "D15", "EGP", BODY, align="center")
put(ws, "E15", "=C15/Assumptions!$B$23", FORMULA, EGP)
put(ws, "F15", "=C15", FORMULA, EGP)
put(ws, "G15", "One-off within the project window", NOTE)

put(ws, "A17", "Sub-total", BOLD, fill=BAND_FILL)
for col in "BCD":
    put(ws, f"{col}17", "", BODY, fill=BAND_FILL)
put(ws, "E17", "=SUM(E5:E15)", BOLD, EGP, BAND_FILL)
put(ws, "F17", "=SUM(F5:F15)", BOLD, EGP, BAND_FILL)
put(ws, "G17", "", BODY, fill=BAND_FILL)

put(ws, "A18", "TOTAL (rounded up)", BOLD, fill=GOOD_FILL)
for col in "BCD":
    put(ws, f"{col}18", "", BODY, fill=GOOD_FILL)
put(ws, "E18", "=CEILING(E17,10)", BOLD, EGP, GOOD_FILL)
put(ws, "F18", "=CEILING(F17,10)", BOLD, EGP, GOOD_FILL)
put(ws, "G18", "Feeds the Cost Breakdown sheet", NOTE, fill=GOOD_FILL)

put(ws, "A20", "PDPL fallback", BOLD, border=False)
c = put(ws, "B20",
        "If legal counsel rejects EU hosting, Egyptian local hosting is estimated at 3,000–8,000 "
        "EGP/month — an increase of roughly 4,000 to 28,500 EGP across the project window, which "
        "must come from contingency. This is an indicative range, not a quote: obtaining two firm "
        "local quotes is a Sprint 0 action item.", NOTE, border=False, fill=WARN_FILL)
c.alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("B20:G22")

# ---------------------------------------------------------------- CASHFLOW ----
ws = wb.create_sheet("Cash Flow")
widths(ws, {"A": 16, "B": 12, "C": 16, "D": 16, "E": 16, "F": 18, "G": 18, "H": 18})
title(ws, "Cash Flow by Month",
      "Labour is spread by working weeks in each month. Infrastructure runs Aug–Dec.")

header(ws, 4, ["Month", "Weeks", "Backend", "Frontend", "DevOps", "Infrastructure",
               "Monthly total", "Cumulative"])

MONTHS = [
    ("Aug 2026", 3.0, "From 10 Aug — Sprint 0 and Sprint 1"),
    ("Sep 2026", 4.3, "Sprints 2–3"),
    ("Oct 2026", 4.4, "Sprints 4–5"),
    ("Nov 2026", 3.3, "Sprint 6, UAT, go-live 20 Nov"),
    ("Dec 2026", 0.0, "Hypercare — infrastructure only"),
]
r = 5
for month, weeks, _ in MONTHS:
    put(ws, f"A{r}", month, BOLD)
    put(ws, f"B{r}", weeks, INPUT, NUM, align="center")
    put(ws, f"C{r}", f"=Assumptions!$B$15*Assumptions!$B$19/Assumptions!$B$10*B{r}", FORMULA, EGP)
    put(ws, f"D{r}", f"=Assumptions!$B$16*Assumptions!$B$20/Assumptions!$B$10*B{r}", FORMULA, EGP)
    put(ws, f"E{r}", f"=Assumptions!$B$17*Assumptions!$B$21/Assumptions!$B$10*B{r}", FORMULA, EGP)
    put(ws, f"F{r}", "='Infra and Tooling'!$F$18/Assumptions!$B$23", LINK, EGP)
    put(ws, f"G{r}", f"=SUM(C{r}:F{r})", FORMULA, EGP)
    put(ws, f"H{r}", f"=G{r}" if r == 5 else f"=H{r-1}+G{r}", FORMULA, EGP)
    r += 1

put(ws, "A10", "Total", BOLD, fill=GOOD_FILL)
put(ws, "B10", "=SUM(B5:B9)", BOLD, NUM, GOOD_FILL, align="center")
for col in "CDEFG":
    put(ws, f"{col}10", f"=SUM({col}5:{col}9)", BOLD, EGP, GOOD_FILL)
put(ws, "H10", "=H9", BOLD, EGP, GOOD_FILL)

put(ws, "A12", "Contingency (undrawn)", BOLD)
put(ws, "B12", "='Cost Breakdown'!$E$13", LINK, EGP)
put(ws, "A13", "Committed + contingency", BOLD)
put(ws, "B13", "=G10+B12", FORMULA, EGP)
put(ws, "A14", "Reconciles to budget", BOLD)
put(ws, "B14", '=IF(ROUND(B13-Assumptions!$B$5,0)=0,"OK","CHECK — see note")', FORMULA, None, GOOD_FILL)

c = put(ws, "D12", "Weeks are working weeks within each calendar month across the 10 Aug – "
                   "20 Nov window. December carries infrastructure only, for hypercare.",
        NOTE, border=False)
c.alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("D12:H14")

# ------------------------------------------------------------ SCOPE CHANGES ----
ws = wb.create_sheet("Scope Changes")
widths(ws, {"A": 46, "B": 11, "C": 11, "D": 11, "E": 15, "F": 54})
title(ws, "Scope Changes — Authentication & Super Admin Console",
      "Added after the first proposal draft. Phase 1 had no slack, so the cost is paid by "
      "displacement rather than absorbed silently.")

put(ws, "A4", "ADDED TO PHASE 1 — cannot wait for peak week", BOLD, fill=BAND_FILL)
for col in "BCDEF":
    put(ws, f"{col}4", "", BODY, fill=BAND_FILL)

header(ws, 5, ["Item", "BE days", "FE days", "Ops days", "Cost (EGP)", "Why it cannot wait"])

ADDED = [
    ("Login, session policy, 2FA (Fortify)", 1.5, 0.0, 0.0,
     "Gates everything else — nothing can be role-scoped until it exists"),
    ("User management + roles UI", 1.0, 0.5, 0.0,
     "Onboarding and offboarding staff during peak week"),
    ("Permission matrix editor", 0.5, 1.0, 0.0,
     "Tuning a role without a deployment"),
    ("Integrations console + credential rotation", 2.0, 1.0, 0.0,
     "Rotating a leaked key, checking webhook health"),
    ("System health + dead-letter replay", 2.0, 0.5, 0.0,
     "Replaying a failed webhook at 2am on Black Friday"),
]
r = 6
for label, be, fe, ops, note in ADDED:
    put(ws, f"A{r}", label, BODY)
    put(ws, f"B{r}", be, INPUT, NUM, align="center")
    put(ws, f"C{r}", fe, INPUT, NUM, align="center")
    put(ws, f"D{r}", ops, INPUT, NUM, align="center")
    put(ws, f"E{r}", f"=B{r}*Assumptions!$B$15/21+C{r}*Assumptions!$B$16/21+D{r}*Assumptions!$B$17/21",
        FORMULA, EGP)
    c = put(ws, f"F{r}", note, NOTE)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    r += 1

put(ws, "A11", "Added to Phase 1", BOLD, fill=WARN_FILL)
for col in "BCD":
    put(ws, f"{col}11", f"=SUM({col}6:{col}10)", BOLD, NUM, WARN_FILL, align="center")
put(ws, "E11", "=SUM(E6:E10)", BOLD, EGP, WARN_FILL)
put(ws, "F11", "Total engineering days added inside the Phase 1 window", NOTE, fill=WARN_FILL)

put(ws, "A13", "DISPLACED TO PHASE 1b — pays for the above", BOLD, fill=BAND_FILL)
for col in "BCDEF":
    put(ws, f"{col}13", "", BODY, fill=BAND_FILL)

header(ws, 14, ["Item", "BE days", "FE days", "Ops days", "Cost (EGP)", "Why it is safe to move"])

MOVED = [
    ("Basic vendor scorecard (§10)", 2.0, 2.0, 0.0,
     "A read model over data captured from day one — ships with full history"),
    ("Executive dashboard (§11)", 1.0, 2.0, 0.0,
     "Same: a view over data already being collected, nothing lost retroactively"),
    ("Vendor product submission (§8.1)", 2.0, 2.0, 0.0,
     "Vendors are onboarded through Ops today, so the manual path already exists"),
]
r = 15
for label, be, fe, ops, note in MOVED:
    put(ws, f"A{r}", label, BODY)
    put(ws, f"B{r}", be, INPUT, NUM, align="center")
    put(ws, f"C{r}", fe, INPUT, NUM, align="center")
    put(ws, f"D{r}", ops, INPUT, NUM, align="center")
    put(ws, f"E{r}", f"=B{r}*Assumptions!$B$15/21+C{r}*Assumptions!$B$16/21+D{r}*Assumptions!$B$17/21",
        FORMULA, EGP)
    c = put(ws, f"F{r}", note, NOTE)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    r += 1

put(ws, "A18", "Displaced out of Phase 1", BOLD, fill=GOOD_FILL)
for col in "BCD":
    put(ws, f"{col}18", f"=SUM({col}15:{col}17)", BOLD, NUM, GOOD_FILL, align="center")
put(ws, "E18", "=SUM(E15:E17)", BOLD, EGP, GOOD_FILL)
put(ws, "F18", "Total engineering days moved to Phase 1b", NOTE, fill=GOOD_FILL)

put(ws, "A20", "NET EFFECT ON PHASE 1", BOLD, fill=HEAD_FILL)
ws["A20"].font = H2
for col in "BCD":
    put(ws, f"{col}20", f"=B11-B18" if col == "B" else f"={col}11-{col}18", BOLD, NUM, HEAD_FILL, align="center")
c = put(ws, "E20", "=E11-E18", BOLD, EGP, HEAD_FILL)
c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
put(ws, "F20", "", BODY, fill=HEAD_FILL)

put(ws, "A22", "Phase 1 budget impact", BOLD)
put(ws, "B22", '=IF(E20<=0,"NONE — the trade pays for itself","SHORTFALL — see Sensitivity")',
    FORMULA, None, GOOD_FILL)
ws.merge_cells("B22:D22")

c = put(ws, "A24",
        "Stated plainly: the console is real work and Phase 1 had no slack, so something had to "
        "give. Displacing three read-only surfaces is the cheapest way to pay for it — they are "
        "views over data the system collects from day one, so they lose no history and simply "
        "arrive three weeks after go-live. The alternative, drawing on contingency, is modelled "
        "on the Sensitivity sheet and is not recommended while the PDPL hosting risk is open.",
        NOTE, border=False)
c.alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("A24:F27")

# ------------------------------------------------------------ PHASE 1B / 2 ----
ws = wb.create_sheet("Phase 1b and 2")
widths(ws, {"A": 40, "B": 12, "C": 12, "D": 12, "E": 16, "F": 46})
title(ws, "Phase 1b and Phase 2 — Indicative Estimates",
      "Not funded by the 300,000 EGP Phase 1 budget. Presented so the client can plan.")

put(ws, "A4", "PHASE 1b — 23 Nov to 11 Dec 2026 (3 weeks)", BOLD, fill=BAND_FILL)
for col in "BCDEF":
    put(ws, f"{col}4", "", BODY, fill=BAND_FILL)

header(ws, 5, ["Deliverable", "BE days", "FE days", "Ops days", "Cost (EGP)", "Notes"])

P1B = [
    ("Customer self-service returns portal", 4, 8, 1, "Order number + phone lookup, no account required"),
    ("Returns analytics dashboard", 4, 3, 0, "By reason, vendor and stakeholder"),
    ("E-stebdal parallel-run validation", 2, 1, 1, "Data parity check before January cutover"),
    ("Super Admin: business rules engine UI", 3, 2, 0, "Phase 1 seeds these rules by migration"),
    ("Super Admin: report builder + scheduling", 2, 2, 0, "Phase 1 ships a fixed set of built-in reports"),
    ("Basic vendor scorecard (displaced from P1)", 2, 2, 0, "Read model - ships with full history"),
    ("Executive dashboard (displaced from P1)", 1, 2, 0, "Read model - ships with full history"),
    ("Vendor product submission (displaced from P1)", 2, 2, 0, "Vendors onboard through Ops in the interim"),
]
r = 6
for label, be, fe, ops, note in P1B:
    put(ws, f"A{r}", label, BODY)
    put(ws, f"B{r}", be, INPUT, NUM, align="center")
    put(ws, f"C{r}", fe, INPUT, NUM, align="center")
    put(ws, f"D{r}", ops, INPUT, NUM, align="center")
    put(ws, f"E{r}",
        f"=B{r}*Assumptions!$B$15/21+C{r}*Assumptions!$B$16/21+D{r}*Assumptions!$B$17/21",
        FORMULA, EGP)
    c = put(ws, f"F{r}", note, NOTE)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    r += 1

put(ws, "A14", "Phase 1b total", BOLD, fill=GOOD_FILL)
for col in "BCD":
    put(ws, f"{col}14", f"=SUM({col}6:{col}13)", BOLD, NUM, GOOD_FILL, align="center")
put(ws, "E14", "=SUM(E6:E13)", BOLD, EGP, GOOD_FILL)
put(ws, "F14", "Assumes 21 working days per month", NOTE, fill=GOOD_FILL)

put(ws, "A16", "PHASE 2 — from January 2027", BOLD, fill=BAND_FILL)
for col in "BCDEF":
    put(ws, f"{col}16", "", BODY, fill=BAND_FILL)

header(ws, 17, ["Deliverable", "BE days", "FE days", "Ops days", "Cost (EGP)", "Requirement"])

P2 = [
    ("Raw fabric inventory, BOM and forecasting", 15, 6, 0, "Req 7.2 — largest single Phase 2 item"),
    ("Automated Paymob Send disbursement", 8, 2, 1, "Req 8.6 — human approval gate retained"),
    ("Second courier adapter", 10, 3, 1, "Req 3.1 — per-item courier selection"),
    ("Advanced vendor scorecard and penalties", 10, 6, 0, "Req 10"),
    ("Executive BI dashboard", 8, 6, 0, "Req 11"),
    ("Automated recall / win-back workflows", 7, 4, 0, "The §1.2 revenue case, automated"),
]
r = 18
for label, be, fe, ops, note in P2:
    put(ws, f"A{r}", label, BODY)
    put(ws, f"B{r}", be, INPUT, NUM, align="center")
    put(ws, f"C{r}", fe, INPUT, NUM, align="center")
    put(ws, f"D{r}", ops, INPUT, NUM, align="center")
    put(ws, f"E{r}",
        f"=B{r}*Assumptions!$B$15/21+C{r}*Assumptions!$B$16/21+D{r}*Assumptions!$B$17/21",
        FORMULA, EGP)
    c = put(ws, f"F{r}", note, NOTE)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    r += 1

put(ws, "A24", "Phase 2 total", BOLD, fill=GOOD_FILL)
for col in "BCD":
    put(ws, f"{col}24", f"=SUM({col}18:{col}23)", BOLD, NUM, GOOD_FILL, align="center")
put(ws, "E24", "=SUM(E18:E23)", BOLD, EGP, GOOD_FILL)
put(ws, "F24", "Excludes infrastructure", NOTE, fill=GOOD_FILL)

put(ws, "A26", "Phase 1 + 1b + 2 combined", BOLD, fill=HEAD_FILL)
ws["A26"].font = H2
for col in "BCD":
    put(ws, f"{col}26", "", BODY, fill=HEAD_FILL)
c = put(ws, "E26", "='Cost Breakdown'!$E$15+E14+E24", BOLD, EGP, HEAD_FILL)
c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
put(ws, "F26", "", BODY, fill=HEAD_FILL)

# ------------------------------------------------------------- SENSITIVITY ----
ws = wb.create_sheet("Sensitivity")
widths(ws, {"A": 34, "B": 16, "C": 16, "D": 16, "E": 62})
title(ws, "Sensitivity Analysis",
      "What the budget buys under different assumptions — and the three levers if the schedule slips.")

put(ws, "A4", "BUDGET SENSITIVITY", BOLD, fill=BAND_FILL)
for col in "BCDE":
    put(ws, f"{col}4", "", BODY, fill=BAND_FILL)

header(ws, 5, ["Scenario", "Budget (EGP)", "Person-months", "Δ vs plan", "Scope consequence"])

SCEN = [
    ("Budget −20%", 0.80, "Phase 1 loses settlement automation AND the vendor scorecard. "
                          "Order, shipping, returns and finance survive."),
    ("Budget −10%", 0.90, "Vendor scorecard deferred to Phase 1b. Everything else holds."),
    ("Plan as proposed", 1.00, "Full Phase 1 scope as defined in proposal §3.1. Zero slack."),
    ("Budget +10%", 1.10, "Restores ~3 weeks of slack — the single most valuable addition, "
                          "given Phase 1 currently has none."),
    ("Budget +20%", 1.20, "Pulls the customer self-service returns portal back into Phase 1."),
]
r = 6
for label, mult, note in SCEN:
    fill = GOOD_FILL if mult == 1.00 else None
    put(ws, f"A{r}", label, BOLD if mult == 1.00 else BODY, fill=fill)
    put(ws, f"B{r}", f"=Assumptions!$B$5*{mult}", FORMULA, EGP, fill)
    # Infrastructure is broadly fixed regardless of budget; contingency scales with it.
    put(ws, f"C{r}", f"=(B{r}-'Infra and Tooling'!$F$18-B{r}*'Cost Breakdown'!$F$13)/"
                     f"'Cost Breakdown'!$B$19", FORMULA, NUM, fill, align="center")
    put(ws, f"D{r}", f"=C{r}-Assumptions!$B$26", FORMULA, NUM, fill, align="center")
    c = put(ws, f"E{r}", note, NOTE, fill=fill)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    r += 1

put(ws, "A12", "TEAM MODEL SENSITIVITY", BOLD, fill=BAND_FILL)
for col in "BCDE":
    put(ws, f"{col}12", "", BODY, fill=BAND_FILL)

header(ws, 13, ["Engagement model", "Rate multiple", "Person-months", "Δ vs plan", "Verdict"])

MODELS = [
    ("In-house salaried (assumed)", 1.0, "Phase 1 scope is achievable. This is the model the "
                                         "budget is built on."),
    ("Freelance / contractor", 1.5, "Buys roughly two thirds of the person-months. Phase 1 must "
                                    "shed settlement, scorecard and vendor portal v1."),
    ("Agency / software house", 2.5, "Buys under half. 300,000 EGP funds an MVP — Shopify sync, "
                                     "item splitting and shipping only. Returns and finance "
                                     "would not fit."),
]
r = 14
for label, mult, note in MODELS:
    fill = GOOD_FILL if mult == 1.0 else WARN_FILL
    put(ws, f"A{r}", label, BOLD if mult == 1.0 else BODY, fill=fill)
    put(ws, f"B{r}", mult, INPUT, '0.0"×"', fill, align="center")
    put(ws, f"C{r}", f"=Assumptions!$B$26/B{r}", FORMULA, NUM, fill, align="center")
    put(ws, f"D{r}", f"=C{r}-Assumptions!$B$26", FORMULA, NUM, fill, align="center")
    c = put(ws, f"E{r}", note, NOTE, fill=fill)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    r += 1

put(ws, "A19", "FUNDING THE ADMIN CONSOLE", BOLD, fill=BAND_FILL)
for col in "BCDE":
    put(ws, f"{col}19", "", BODY, fill=BAND_FILL)

header(ws, 20, ["Option", "Cost (EGP)", "Contingency left", "Preference", "Consequence"])

FUND = [
    ("1. Displace 11 days of read-only UI", 0, "recommended",
     "Vendor scorecard, executive dashboard and vendor product submission move to Phase 1b. "
     "All three are views over data captured from day one, so nothing is lost retroactively."),
    ("2. Draw on contingency", None, "not while PDPL is open",
     "Arithmetically it fits, but leaves too little against risk R1 - the PDPL hosting fallback "
     "alone could need up to 28,500 EGP."),
    ("3. Increase the budget", None, "client's call",
     "Roughly 16,000-20,000 EGP. Keeps every Phase 1 surface and restores the contingency."),
]
r = 21
for label, cost, pref, note in FUND:
    fill = GOOD_FILL if r == 21 else WARN_FILL
    put(ws, f"A{r}", label, BOLD if r == 21 else BODY, fill=fill)
    if cost is None:
        put(ws, f"B{r}", "='Scope Changes'!$E$11", LINK, EGP, fill)
        put(ws, f"C{r}", "='Cost Breakdown'!$E$13-B%d" % r, FORMULA, EGP, fill)
    else:
        put(ws, f"B{r}", cost, INPUT, EGP, fill)
        put(ws, f"C{r}", "='Cost Breakdown'!$E$13", LINK, EGP, fill)
    put(ws, f"D{r}", pref, BODY, fill=fill, align="center")
    c = put(ws, f"E{r}", note, NOTE, fill=fill)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    r += 1

put(ws, "A26", "SCHEDULE LEVERS", BOLD, fill=BAND_FILL)
for col in "BCDE":
    put(ws, f"{col}26", "", BODY, fill=BAND_FILL)

header(ws, 27, ["Lever", "Cost (EGP)", "Time recovered", "Preference", "Trade-off"])

LEVERS = [
    ("1. Cut scope", 0, "~2 weeks", "Preferred",
     "Defer settlement automation and vendor scorecard to Phase 1b. No cash cost."),
    ("2. Extend timeline", 0, "Unlimited", "Second",
     "Go live in December, after Black Friday. Removes all peak-week risk but forfeits a full "
     "peak season of benefit."),
    ("3. Add a 2nd backend engineer", None, "~3 weeks", "Least effective",
     "Only works if committed by mid-September. Later than that, onboarding drag exceeds the "
     "capacity added."),
]
r = 28
for label, cost, time, pref, note in LEVERS:
    put(ws, f"A{r}", label, BOLD)
    if cost is None:
        put(ws, f"B{r}", "=Assumptions!$B$15*3*1", FORMULA, EGP)
    else:
        put(ws, f"B{r}", cost, INPUT, EGP)
    put(ws, f"C{r}", time, BODY, align="center")
    put(ws, f"D{r}", pref, BODY, align="center")
    c = put(ws, f"E{r}", note, NOTE)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    r += 1

# --------------------------------------------------------------------- ROI ----
ws = wb.create_sheet("ROI")
widths(ws, {"A": 46, "B": 18, "C": 18, "D": 58})
title(ws, "Return on Investment",
      "The case rests on recovered failed deliveries, not on data-entry savings. Proposal §1.2.")

put(ws, "A4", "OPERATIONAL BASELINE — from the client's own data", BOLD, fill=BAND_FILL)
for col in "BCD":
    put(ws, f"{col}4", "", BODY, fill=BAND_FILL)

header(ws, 5, ["Measure", "Value", "Unit", "Source"])

BASE = [
    ("Item rows analysed", 18341, "items", "Copy of Orders, 1 Jan – 8 Aug 2026"),
    ("Failed delivery (FD Allow Open Shipment)", 3743, "items", "Final status column"),
    ("Returned to origin (RTO)", 2359, "items", "Final status column"),
    ("Total failed items", None, "items", "FD + RTO"),
    ("Average item value", 897, "EGP", "GMV ÷ item rows"),
    ("GMV that shipped and came back", None, "EGP", "Failed items × average value"),
    ("", None, None, None),
    ("Failed items actually worked (Send messages)", 1430, "items", "Recall Status column"),
    ("Recovered into new orders", 273, "items", "Recall Status = Create new orders"),
    ("Current recovery rate", None, "%", "Recovered ÷ worked"),
    ("Failed items never worked", None, "items", "Total failed − worked"),
]
r = 6
for label, value, unit, source in BASE:
    if not label:
        r += 1
        continue
    put(ws, f"A{r}", label, BODY)
    formulas = {9: "=B7+B8", 11: "=B9*B10", 15: "=B14/B13", 16: "=B9-B13"}
    if value is None:
        put(ws, f"B{r}", formulas[r], FORMULA, PCT if r == 15 else EGP if r == 11 else '#,##0')
    else:
        put(ws, f"B{r}", value, INPUT, EGP if r == 10 else '#,##0')
    put(ws, f"C{r}", unit, BODY, align="center")
    c = put(ws, f"D{r}", source, NOTE)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    r += 1

put(ws, "A18", "RECOVERY SCENARIOS — annualised", BOLD, fill=BAND_FILL)
for col in "BCD":
    put(ws, f"{col}18", "", BODY, fill=BAND_FILL)

header(ws, 19, ["Scenario", "Recovered items / yr", "Recovered GMV / yr (EGP)", "Assumption"])

put(ws, "A20", "Conservative", BODY, fill=GOOD_FILL)
put(ws, "B20", "=B16*(B15/2)*12/7", FORMULA, '#,##0', GOOD_FILL, align="center")
put(ws, "C20", "=B20*B10", FORMULA, EGP, GOOD_FILL)
c = put(ws, "D20", "Half the current recovery rate applied to items never worked. Deliberately "
                   "pessimistic: worked items were probably the most promising.", NOTE, fill=GOOD_FILL)
c.alignment = Alignment(wrap_text=True, vertical="center")

put(ws, "A21", "At parity", BODY)
put(ws, "B21", "=(B9*B15-B14)*12/7", FORMULA, '#,##0', align="center")
put(ws, "C21", "=B21*B10", FORMULA, EGP)
c = put(ws, "D21", "Current 19.1% recovery rate applied across all failed items, net of what is "
                   "already recovered today.", NOTE)
c.alignment = Alignment(wrap_text=True, vertical="center")

put(ws, "A23", "PAYBACK", BOLD, fill=BAND_FILL)
for col in "BCD":
    put(ws, f"{col}23", "", BODY, fill=BAND_FILL)

put(ws, "A24", "Phase 1 build cost", BODY)
put(ws, "B24", "='Cost Breakdown'!$E$15", LINK, EGP)
put(ws, "A25", "Recovered GMV, conservative case", BODY)
put(ws, "B25", "=C20", FORMULA, EGP)
put(ws, "A26", "Ratio — recovered GMV to build cost", BOLD, fill=GOOD_FILL)
put(ws, "B26", "=B25/B24", BOLD, '0.0"×"', GOOD_FILL, align="center")

c = put(ws, "D24", "These are GMV figures, not profit. The platform realises commission and "
                   "margin on recovered orders, not the full value. The point is the order of "
                   "magnitude: even the conservative case returns more than twice the build cost "
                   "in GMV within year one, and the recall queue is a permanent capability.",
        NOTE, border=False)
c.alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("D24:D27")

put(ws, "A28", "RECURRING OFFSETS", BOLD, fill=BAND_FILL)
for col in "BCD":
    put(ws, f"{col}28", "", BODY, fill=BAND_FILL)

put(ws, "A29", "E-stebdal subscription — monthly", BODY, fill=FLAG_FILL)
put(ws, "B29", 0, INPUT, EGP, FLAG_FILL, align="center")
c = put(ws, "D29", "CLIENT TO CONFIRM. Enter the current monthly subscription; the annual saving "
                   "below will calculate. Retired at January 2027 cutover.", NOTE, fill=FLAG_FILL)
c.alignment = Alignment(wrap_text=True, vertical="center")

put(ws, "A30", "Annual saving from retiring E-stebdal", BOLD)
put(ws, "B30", "=B29*12", FORMULA, EGP, align="center")

put(ws, "A31", "Manual data entry displaced", BODY)
put(ws, "B31", "=B6/155", FORMULA, '#,##0', align="center")
put(ws, "C31", "rows / working day", BODY)
c = put(ws, "D31", "Across 32 columns with cross-system lookups per row. A real but soft saving — "
                   "not counted in the payback figure above.", NOTE)
c.alignment = Alignment(wrap_text=True, vertical="center")

for sheet in wb.worksheets:
    sheet.sheet_view.showGridLines = False

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"wrote {OUT.relative_to(ROOT)}")
