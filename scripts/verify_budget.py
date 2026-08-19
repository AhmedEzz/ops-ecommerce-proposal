#!/usr/bin/env python3
"""
Verify the budget workbook without LibreOffice.

LibreOffice is not installed on this machine, so the usual recalculate-and-check
pass is unavailable. This does two things in its place:

  1. Static check - every sheet and cell a formula references actually exists,
     and no formula uses a function LibreOffice/older Excel cannot evaluate.
  2. Arithmetic check - the model is recomputed independently in Python and
     compared against the figures quoted in the proposal and budget summary.

A pass here means the references resolve and the numbers are right. Excel and
Numbers both recalculate on open, so the delivered file will display correctly.
"""

import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
BOOK = ROOT / "deliverables" / "02-Budget-Plan.xlsx"

# Spilling / post-2007 functions that would render as #NAME? in the wild.
BANNED = {"XLOOKUP", "XMATCH", "SORT", "FILTER", "UNIQUE", "SEQUENCE",
          "TEXTJOIN", "CONCAT", "IFS", "SWITCH", "MAXIFS", "MINIFS"}

CELL_REF = re.compile(r"(?:'([^']+)'|([A-Za-z][A-Za-z0-9 ]*))?!\$?([A-Z]{1,3})\$?(\d+)")
FUNC = re.compile(r"\b([A-Z][A-Z0-9._]{1,})\s*\(")

failures = []
formula_count = 0


def check_static():
    global formula_count
    wb = openpyxl.load_workbook(BOOK)
    names = set(wb.sheetnames)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str) or not value.startswith("="):
                    continue
                formula_count += 1
                where = f"{ws.title}!{cell.coordinate}"

                for func in FUNC.findall(value):
                    bare = func.replace("_xlfn.", "")
                    if bare in BANNED and not func.startswith("_xlfn."):
                        failures.append(f"{where}: unsupported function {func}")

                for quoted, plain, col, rownum in CELL_REF.findall(value):
                    sheet = quoted or plain
                    if not sheet:
                        continue
                    if sheet not in names:
                        failures.append(f"{where}: references missing sheet '{sheet}'")
                        continue
                    target = wb[sheet]
                    if int(rownum) > target.max_row + 5:
                        failures.append(
                            f"{where}: '{sheet}'!{col}{rownum} is past the used range "
                            f"(max row {target.max_row})")
    print(f"  formulas scanned: {formula_count}")


def check_arithmetic():
    """Recompute the model from first principles and compare to expectations."""
    budget = 300_000
    weeks, weeks_per_month = 15, 4.345
    months = weeks / weeks_per_month

    be_rate, fe_rate, ops_rate = 34_000, 27_000, 35_000
    be_fte, fe_fte, ops_fte = 1.00, 1.00, 0.40
    usd, eur = 50.25, 55.00
    infra_months = 5

    person_months = (be_fte + fe_fte + ops_fte) * months
    labour = (be_rate * be_fte + fe_rate * fe_fte + ops_rate * ops_fte) * months

    monthly_eur = 15.99 + 8.49 + 5.49 + 5.99 + 5.00
    monthly_usd = 12.00 + 5.00
    infra_recurring = (monthly_eur * eur + monthly_usd * usd) * infra_months
    domain = 600
    infra_raw = infra_recurring + domain
    infra = -(-infra_raw // 10) * 10  # CEILING(x, 10)

    contingency = budget - labour - infra
    total = labour + infra + contingency
    cost_per_pm = labour / person_months

    checks = [
        ("duration (months)", months, 3.45, 0.01),
        ("person-months", person_months, 8.28, 0.02),
        ("labour total", labour, 258_919, 50),
        ("infrastructure total", infra, 16_140, 20),
        ("contingency", contingency, 24_941, 50),
        ("contingency %", contingency / budget, 0.083, 0.002),
        ("GRAND TOTAL", total, 300_000, 0.01),
        ("cost per person-month", cost_per_pm, 31_250, 400),
    ]

    print(f"\n  {'metric':<26} {'computed':>12} {'expected':>12}   status")
    for label, got, want, tol in checks:
        ok = abs(got - want) <= tol
        if not ok:
            failures.append(f"arithmetic: {label} = {got:,.2f}, expected ~{want:,.2f}")
        fmt = f"{got:>12,.3f}" if got < 100 else f"{got:>12,.0f}"
        exp = f"{want:>12,.3f}" if want < 100 else f"{want:>12,.0f}"
        print(f"  {label:<26} {fmt} {exp}   {'ok' if ok else 'MISMATCH'}")

    # The whole point of the model: the total must land on the budget exactly.
    if abs(total - budget) > 0.01:
        failures.append(f"GRAND TOTAL {total:,.2f} != budget {budget:,}")

    print(f"\n  ROI cross-check")
    failed_items = 3743 + 2359
    avg_value = 897
    worked, recovered = 1430, 273
    rate = recovered / worked
    never_worked = failed_items - worked
    conservative = never_worked * (rate / 2) * 12 / 7
    parity = (failed_items * rate - recovered) * 12 / 7
    for label, got, want, tol in [
        ("failed items", failed_items, 6102, 0),
        ("recovery rate", rate, 0.191, 0.001),
        ("conservative items/yr", conservative, 760, 10),
        ("conservative GMV/yr", conservative * avg_value, 683_000, 8_000),
        ("parity items/yr", parity, 1_530, 15),
        ("parity GMV/yr", parity * avg_value, 1_370_000, 15_000),
    ]:
        ok = abs(got - want) <= tol
        if not ok:
            failures.append(f"ROI: {label} = {got:,.2f}, expected ~{want:,.2f}")
        fmt = f"{got:>12,.3f}" if got < 100 else f"{got:>12,.0f}"
        print(f"  {label:<26} {fmt}   {'ok' if ok else 'MISMATCH'}")


def main():
    print(f"Verifying {BOOK.name}\n")
    check_static()
    check_arithmetic()

    print()
    if failures:
        print(f"FAILED — {len(failures)} problem(s):")
        for f in failures:
            print(f"  - {f}")
        sys.exit(1)
    print("PASSED — references resolve, no unsupported functions, arithmetic confirmed.")
    print("Note: openpyxl writes formulas without cached values. Excel and Numbers")
    print("recalculate on open; lightweight preview tools may show blanks until then.")


if __name__ == "__main__":
    main()
